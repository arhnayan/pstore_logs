#!/usr/bin/env python3
"""
Vodafone Storage Performance Report Generator
Generates formatted Excel reports with charts and anomaly detection
"""

import os
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule, FormulaRule
from datetime import datetime
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

# Define locations and their servers (default seed).
DEFAULT_LOCATION_SERVERS = {
    'Adana': [
        'ADNTOCDELLCLD01', 'ADNTOCDELLAPP01', 'ADNTOCDELLAPP02',
        'ADNTOCDELLAPP03', 'ADNTOCDELLAPP04', 'ADNTOCDELLOSS01'
    ],
    'Diyarbakır': ['DYBTOCDELLAPP01'],
    'Esenyurt': [
        'ESNTOCDELLCLD01', 'ESNTOCDELLAPP01', 'ESNTOCDELLAPP02',
        'ESNTOCDELLAPP03', 'ESNTOCDELLAPP04', 'ESNTOCDELLOSS01'
    ],
    'Gaziemir': [
        'IZMTOCDELLCLD01', 'IZMTOCDELLAPP01', 'IZMTOCDELLAPP02',
        'IZMTOCDELLAPP03', 'IZMTOCDELLAPP04', 'IZMTOCDELLOSS01'
    ],
    'Pursaklar': [
        'ANKTOCDELLCLD01', 'ANKTOCDELLAPP01', 'ANKTOCDELLAPP02',
        'ANKTOCDELLAPP03', 'ANKTOCDELLAPP04', 'ANKTOCDELLOSS01'
    ],
    'Tuzla': [
        'TZLTOCDELLCLD01', 'TZLTOCDELLAPP01', 'TZLTOCDELLAPP02',
        'TZLTOCDELLAPP03', 'TZLTOCDELLAPP04', 'TZLTOCDELLOSS01'
    ]
}

class ReportGenerator:
    def __init__(
        self,
        raw_csv_dir='raw_report_csv',
        output_dir='reports',
        capacity_data=None,
        formatted_csv_dir=None,
        load_capacity_from_formatted_csv=False,
        enable_analytics=True,
        location_servers=None,
        server_data=None,
    ):
        self.raw_csv_dir = raw_csv_dir
        self.location_servers = location_servers or DEFAULT_LOCATION_SERVERS
        self._server_data = server_data or {}
        self._capacity_data: dict[str, dict] = {}
        # Kept for backward compatibility (optional)
        self.formatted_csv_dir = formatted_csv_dir
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.enable_analytics = enable_analytics
        
        # Color definitions
        self.RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        self.BLACK_FILL = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        self.GREY_FILL = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        self.YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.LIGHT_RED_FILL = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        self.WHITE_FONT = Font(color='FFFFFF', bold=True)
        self.BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Capacity is optional. Prefer manual capacity_data (GUI) unless explicitly
        # told to load from formatted CSVs.
        if capacity_data:
            self.set_capacity_data(capacity_data)
        elif load_capacity_from_formatted_csv and self.formatted_csv_dir:
            self.load_capacity_data()

    def set_capacity_data(self, capacity_data):
        """Set capacity data used when writing reports."""
        if not isinstance(capacity_data, dict):
            raise TypeError("capacity_data must be a dict like {SERVER: {'Total_TB':..., 'Free_TB':..., ...}}")
        normalized = {}
        for server, data in capacity_data.items():
            if not server:
                continue
            normalized[str(server).strip().upper()] = dict(data or {})
        self._capacity_data = normalized
        
    def load_capacity_data(self):
        """Load capacity data from formatted CSV files."""
        if not self.formatted_csv_dir or not os.path.exists(self.formatted_csv_dir):
            print(f"Warning: {self.formatted_csv_dir} not found. Using empty capacity data.")
            return

        for filename in os.listdir(self.formatted_csv_dir):
            if not filename.endswith('.csv') or filename == 'Cover-Table 1.csv':
                continue
            
            filepath = os.path.join(self.formatted_csv_dir, filename)
            try:
                # Read the formatted CSV - it has a specific structure
                with open(filepath, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                
                current_server = None
                for i, line in enumerate(lines):
                    line = line.strip()
                    
                    # Check if this is a server name line (uppercase, starts with location code)
                    if line and not line.startswith('CAPACITY') and not line.startswith('Total') and ';' in line:
                        parts = line.split(';')
                        if parts[0] and parts[0].isupper() and 'DELL' in parts[0]:
                            current_server = parts[0]
                        
                        # Check if this is a data line with capacity info
                        if current_server and len(parts) >= 3:
                            try:
                                # Try to parse capacity values
                                total_val = parts[0].replace(',', '.').strip()
                                free_val = parts[1].replace(',', '.').strip()
                                used_val = parts[2].replace(',', '.').strip()
                                
                                if total_val and free_val and total_val.replace('.', '').isdigit():
                                    total_tb = float(total_val)
                                    free_tb = float(free_val)
                                    
                                    # Check if used is in GB or TB
                                    if used_val:
                                        used_num = float(used_val)
                                        if used_num < 10:  # Likely TB
                                            self._capacity_data[current_server] = {
                                                'Total_TB': total_tb,
                                                'Free_TB': free_tb,
                                                'Used_TB': used_num
                                            }
                                        else:  # Likely GB
                                            self._capacity_data[current_server] = {
                                                'Total_TB': total_tb,
                                                'Free_TB': free_tb,
                                                'Used_GB': used_num
                                            }
                                    current_server = None  # Reset after parsing data
                            except (ValueError, IndexError):
                                pass
            except Exception as e:
                print(f"Warning: Could not parse {filename}: {e}")
        
        print(f"Loaded capacity data for {len(self._capacity_data)} servers")
    
    def parse_csv_filename(self, filename):
        """Extract location and server name from filename."""
        match = re.match(r'(.+?)_(.+?)-Table', filename)
        if match:
            location = match.group(1)
            server = match.group(2)
            return location, server
        return None, None
    
    def read_performance_data(self, filepath):
        """Read and parse performance data from CSV."""
        try:
            # Read first line to get column names
            with open(filepath, 'r', encoding='utf-8') as f:
                first_line = f.readline().strip()
            
            # Split the first line by semicolon to get column names
            col_names = first_line.split(';')
            col_names = [col.strip() for col in col_names]
            
            # Replace first column name (server name) with 'Timestamp'
            if len(col_names) > 0:
                col_names[0] = 'Timestamp'
            
            # Now read the CSV data, skipping the first row (header)
            df = pd.read_csv(filepath, sep=';', skiprows=1, names=col_names, encoding='utf-8')
            
            if df.shape[0] < 1:
                return None
            
            original_rows = df.shape[0]
            
            # Convert European decimal format (comma) to standard format for all data columns
            for col in df.columns:
                if col != 'Timestamp':
                    df[col] = df[col].astype(str).str.replace(',', '.').str.replace('%', '').str.strip()
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Get data columns (excluding Timestamp)
            data_cols = [col for col in df.columns if col != 'Timestamp']
            
            if len(data_cols) == 0:
                return None
            
            # Remove rows with all NaN values in data columns
            df = df.dropna(how='all', subset=data_cols)
            
            # Remove rows where all numeric values are 0 (no real data)
            df = df[~((df[data_cols] == 0).all(axis=1))]
            
            # Remove rows where key metrics are missing or invalid
            # A valid row should have at least Total IOPS or CPU Utilization with meaningful values
            key_metrics = [col for col in ['Total IOPS', 'CPU Utilization', 'Latency'] if col in df.columns]
            
            if len(key_metrics) > 0:
                # Keep rows where at least one key metric has a non-zero, non-null value
                mask = pd.Series([False] * len(df), index=df.index)
                for metric in key_metrics:
                    mask = mask | ((df[metric].notna()) & (df[metric] > 0))
                df = df[mask]
            
            # Additional cleanup: remove rows with empty timestamp
            df = df[df['Timestamp'].notna()]
            df = df[df['Timestamp'].astype(str).str.strip() != '']
            
            # Reset index after filtering
            df = df.reset_index(drop=True)
            
            if df.shape[0] > 0:
                return df
            
        except Exception as e:
            print(f"Error reading {filepath}: {e}")
            import traceback
            traceback.print_exc()
        
        return None
    
    def calculate_statistics(self, df, column):
        """Calculate statistics for anomaly detection."""
        values = df[column].dropna()
        
        # Filter out zero values for more accurate statistics
        values = values[values > 0]
        
        if len(values) == 0:
            return {}
        
        # Calculate IQR for outlier detection
        q1 = values.quantile(0.25)
        q3 = values.quantile(0.75)
        iqr = q3 - q1
        
        # Optional: Remove extreme outliers (beyond 3*IQR) for cleaner statistics
        # This helps when there are a few extreme spikes
        lower_bound = q1 - 3 * iqr
        upper_bound = q3 + 3 * iqr
        filtered_values = values[(values >= lower_bound) & (values <= upper_bound)]
        
        # Use filtered values for mean/median if we still have enough data
        if len(filtered_values) > len(values) * 0.8:  # Only if we're not filtering too much (keep 80%+)
            calc_values = filtered_values
        else:
            calc_values = values
        
        return {
            'mean': calc_values.mean(),
            'median': values.median(),  # Use unfiltered for median
            'std': values.std(),
            'min': values.min(),
            'max': values.max(),
            'q1': q1,
            'q3': q3,
            'count': len(values)
        }
    
    def detect_anomalies(self, df, column, threshold_std=2, detect_drops=True, drop_threshold=1.5):
        """Detect anomalies using z-score method. Detects both spikes and drops."""
        values = df[column].dropna()
        
        # Filter out zeros for better anomaly detection
        values = values[values > 0]
        
        if len(values) < 3:
            return []
        
        mean = values.mean()
        std = values.std()
        
        if std == 0:
            return []
        
        # Calculate z-scores (how many standard deviations away from mean)
        z_scores = (values - mean) / std
        
        anomaly_indices = []
        
        # Detect high anomalies (spikes) - values significantly above mean
        high_anomalies = values[z_scores > threshold_std].index.tolist()
        anomaly_indices.extend(high_anomalies)
        
        # Detect low anomalies (drops) - values significantly below mean
        # Use a more sensitive threshold for drops as they're often more concerning
        if detect_drops:
            low_anomalies = values[z_scores < -drop_threshold].index.tolist()
            anomaly_indices.extend(low_anomalies)
        
        return anomaly_indices
    
    def create_header(self, ws, location, servers, row=1):
        """Create the red header with location and server names."""
        server_list = ', '.join(servers)
        header_text = f"{location} Site  {server_list}"
        
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1)
        cell.value = header_text
        cell.fill = self.RED_FILL
        cell.font = self.WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        return row + 2  # Skip one row
    
    def create_server_section(self, ws, server_name, capacity_data, perf_stats, start_row):
        """Create a section for one server with capacity and performance data."""
        current_row = start_row
        
        # Server name header (black)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        cell = ws.cell(row=current_row, column=1)
        cell.value = server_name
        cell.fill = self.BLACK_FILL
        cell.font = self.WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Grey subheaders for CAPACITY and PERFORMANCE
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cap_cell = ws.cell(row=current_row, column=1)
        cap_cell.value = "CAPACITY"
        cap_cell.fill = self.GREY_FILL
        cap_cell.font = Font(bold=True)
        cap_cell.alignment = Alignment(horizontal='center', vertical='center')
        cap_cell.border = self.BORDER
        
        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=7)
        perf_cell = ws.cell(row=current_row, column=4)
        perf_cell.value = "PERFORMANCE"
        perf_cell.fill = self.GREY_FILL
        perf_cell.font = Font(bold=True)
        perf_cell.alignment = Alignment(horizontal='center', vertical='center')
        perf_cell.border = self.BORDER
        current_row += 1
        
        # Column headers
        headers = ['Total (TB)', 'Free (TB)', 'Used (TB)', 'Latency', 'I/O Size', 'Total IOPS', 'CPU Utilization']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = self.GREY_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
        current_row += 1
        
        # Data row
        # Capacity data
        total_tb = capacity_data.get('Total_TB', None)
        free_tb = capacity_data.get('Free_TB', None)
        
        # Calculate used in TB
        if 'Used_TB' in capacity_data:
            used_tb = capacity_data['Used_TB']
        elif 'Used_GB' in capacity_data:
            used_tb = capacity_data['Used_GB'] / 1000  # Convert GB to TB
        else:
            used_tb = None
        
        # Write numeric cells (so formulas/conditional formatting can work after manual edits)
        total_cell = ws.cell(row=current_row, column=1)
        free_cell = ws.cell(row=current_row, column=2)
        used_cell = ws.cell(row=current_row, column=3)

        total_cell.value = float(total_tb) if total_tb not in (None, '') else None
        free_cell.value = float(free_tb) if free_tb not in (None, '') else None

        # Always link Used to Total/Free so manual edits in Excel remain consistent,
        # even when no capacity_data was provided at generation time.
        used_cell.value = f'=IF(OR(A{current_row}="",B{current_row}=""),"",A{current_row}-B{current_row})'

        for c in (total_cell, free_cell, used_cell):
            c.number_format = '0.00'
        
        # Capacity utilization bar (formula-driven, so it updates after manual edits)
        # This sits directly below the capacity columns (A-C) for the server.
        gauge_row = current_row + 1

        # Label
        label_cell = ws.cell(row=gauge_row, column=1)
        label_cell.value = "Capacity"
        label_cell.font = Font(bold=True, italic=True)
        label_cell.alignment = Alignment(horizontal='left')
        label_cell.border = self.BORDER

        # Usage % cell (with formula referencing Total/Free cells above).
        # Interpreted as a fraction (0–1) so that higher usage = more bar fill.
        # If Total or Free is empty, keep this cell blank.
        bar_cell = ws.cell(row=gauge_row, column=2)
        bar_cell.value = f'=IF(OR(A{current_row}="",B{current_row}=""),"", (A{current_row}-B{current_row})/A{current_row})'
        bar_cell.number_format = '0.0%'
        bar_cell.border = self.BORDER

        # Data Bar (single rule, value comes from formula).
        # Underlying value is a fraction (0–1), so max is 1.0 → fully filled bar.
        data_bar = DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=1,
            color="63BE7B",  # base green bar
            showValue=True,
            minLength=0,
            maxLength=100
        )
        ws.conditional_formatting.add(f"B{gauge_row}:B{gauge_row}", data_bar)

        # Percentage text cell (mirrors bar cell, easier to read)
        pct_cell = ws.cell(row=gauge_row, column=3)
        pct_cell.value = f'=IF(B{gauge_row}="","",TEXT(B{gauge_row},"0.0%"))'
        pct_cell.font = Font(bold=True)
        pct_cell.alignment = Alignment(horizontal="center")
        pct_cell.border = self.BORDER
        
        # Performance stats (averages)
        latency_avg = perf_stats.get('Latency', {}).get('mean', None)
        io_size_avg = perf_stats.get('Avg. Size', {}).get('mean', None)
        iops_avg = perf_stats.get('Total IOPS', {}).get('mean', None)
        iops_max = perf_stats.get('Total IOPS', {}).get('max', None)
        cpu_avg = perf_stats.get('CPU Utilization', {}).get('mean', None)
        
        # Write performance data
        if latency_avg is not None and latency_avg > 0:
            ws.cell(row=current_row, column=4).value = f"{latency_avg:.2f}"
        if io_size_avg is not None and io_size_avg > 0:
            ws.cell(row=current_row, column=5).value = f"{io_size_avg:.0f}"
        if iops_avg is not None and iops_avg > 0:
            ws.cell(row=current_row, column=6).value = f"{iops_avg:.0f}"
        if cpu_avg is not None and cpu_avg > 0:
            ws.cell(row=current_row, column=7).value = f"{cpu_avg:.1f}%"
        
        # Add conditional formatting for performance warnings
        if cpu_avg and cpu_avg > 0:
            if cpu_avg > 80:  # High CPU usage - RED
                ws.cell(row=current_row, column=7).fill = self.LIGHT_RED_FILL
            elif cpu_avg > 60:  # Moderate CPU usage - YELLOW
                ws.cell(row=current_row, column=7).fill = self.YELLOW_FILL
        
        if iops_max and iops_max > 25000:  # High IOPS detected
            ws.cell(row=current_row, column=6).fill = self.YELLOW_FILL
        
        # Apply borders to data cells
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = self.BORDER
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        return current_row + 2  # Skip one row after each server
    
    def create_charts_sheet(self, wb, location, all_data):
        """Create a sheet with charts for anomaly detection and analysis."""
        ws = wb.create_sheet(f"{location} - Analytics")
        
        row = 1
        ws.cell(row=row, column=1).value = f"{location} - Performance Analytics & Anomaly Detection"
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        row += 2
        
        # Write summary statistics
        ws.cell(row=row, column=1).value = "Özet İstatistikler ve Anormallik Tespiti (Tam Veri Seti)"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        row += 1
        
        ws.cell(row=row, column=1).value = "Not: Anormallik sayıları tam veri setinden (~720 veri noktası) alınmıştır. Aşağıdaki grafikler örneklenmiş verileri (50 nokta) göstermektedir."
        ws.cell(row=row, column=1).font = Font(italic=True, size=9, color='666666')
        row += 1
        
        headers = ['Server', 'Metric', 'Mean', 'Median', 'Std Dev', 'Min', 'Max', 'Anomalies']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = self.GREY_FILL
            cell.border = self.BORDER
        stats_start_row = row + 1
        row += 1
        
        # Calculate and write statistics
        anomaly_details = []
        for server_name, df in all_data.items():
            if df is None or df.empty:
                continue
            
            for metric in ['CPU Utilization', 'Total IOPS', 'Latency']:
                if metric in df.columns:
                    stats = self.calculate_statistics(df, metric)
                    anomalies = self.detect_anomalies(df, metric)
                    
                    if stats:
                        ws.cell(row=row, column=1).value = server_name
                        ws.cell(row=row, column=2).value = metric
                        ws.cell(row=row, column=3).value = round(stats['mean'], 2)
                        ws.cell(row=row, column=4).value = round(stats['median'], 2)
                        ws.cell(row=row, column=5).value = round(stats['std'], 2)
                        ws.cell(row=row, column=6).value = round(stats['min'], 2)
                        ws.cell(row=row, column=7).value = round(stats['max'], 2)
                        
                        # Show anomaly count with percentage
                        total_points = stats.get('count', len(df))
                        anomaly_count = len(anomalies)
                        if total_points > 0:
                            anomaly_pct = (anomaly_count / total_points) * 100
                            ws.cell(row=row, column=8).value = f"{anomaly_count} ({anomaly_pct:.1f}%)"
                        else:
                            ws.cell(row=row, column=8).value = anomaly_count
                        
                        # Highlight if anomalies detected
                        if len(anomalies) > 0:
                            ws.cell(row=row, column=8).fill = self.YELLOW_FILL
                            anomaly_details.append(f"{server_name} - {metric}: {len(anomalies)} anomalies")
                        
                        # Highlight high CPU or IOPS
                        if metric == 'CPU Utilization' and stats['mean'] > 60:
                            ws.cell(row=row, column=3).fill = self.LIGHT_RED_FILL
                        if metric == 'Total IOPS' and stats['max'] > 25000:
                            ws.cell(row=row, column=7).fill = self.LIGHT_RED_FILL
                        
                        for col in range(1, 9):
                            ws.cell(row=row, column=col).border = self.BORDER
                        
                        row += 1
        
        stats_end_row = row - 1
        
        # Add charts section
        row += 2
        ws.cell(row=row, column=1).value = "Görsel Analiz"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        row += 1
        
        # Add legend explanation
        ws.cell(row=row, column=1).value = "Açıklama: Her cihazın anormallikleri farklı şekil/renk işaretlerle gösterilmiştir (▼◆●■ vb.)"
        ws.cell(row=row, column=1).font = Font(italic=True, size=10)
        ws.cell(row=row, column=1).fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        row += 1
        
        ws.cell(row=row, column=1).value = "Anormallik kriteri: Spikes >2σ OR Drops <-1.5σ (ortalamadan standart sapma)"
        ws.cell(row=row, column=1).font = Font(italic=True, size=9, color='666666')
        row += 1
        
        ws.cell(row=row, column=1).value = "Not: Grafikler her 10. veri noktasını gösterir (720 toplam noktadan 50 nokta örneklenmiştir). Tüm anormallikler görünür olmayabilir."
        ws.cell(row=row, column=1).font = Font(italic=True, size=9, color='666666')
        row += 1
        
        # Create CPU Utilization chart data
        chart_start_row = row
        ws.cell(row=row, column=1).value = "Time Interval"
        col_idx = 2
        server_cols = {}
        
        # First, add all server data columns
        for server_name, df in all_data.items():
            if df is not None and not df.empty and 'CPU Utilization' in df.columns:
                ws.cell(row=row, column=col_idx).value = f"{server_name} CPU%"
                server_cols[server_name] = col_idx
                col_idx += 1
        
        # Add anomaly marker columns (vertical lines)
        anomaly_line_start_col = col_idx
        ws.cell(row=row, column=col_idx).value = "Anomaly Marker"
        col_idx += 1
        
        # Collect all anomalies across servers
        all_anomaly_timestamps = set()
        server_anomalies = {}
        
        for server_name, df in all_data.items():
            if df is not None and not df.empty and 'CPU Utilization' in df.columns:
                anomaly_indices = self.detect_anomalies(df, 'CPU Utilization')
                server_anomalies[server_name] = anomaly_indices
                all_anomaly_timestamps.update(anomaly_indices)
        
        # Write sample data (every 10th point to avoid cluttering)
        data_row = row + 1
        max_cpu_value = 0
        
        # First pass - write data and find max value
        for server_name, df in all_data.items():
            if df is not None and not df.empty and 'CPU Utilization' in df.columns and server_name in server_cols:
                cpu_data = df['CPU Utilization'].dropna()
                max_cpu_value = max(max_cpu_value, cpu_data.max())
                
                # Sample every 10th point, max 50 points
                sample_indices = range(0, min(len(cpu_data), 500), 10)
                
                for idx, data_idx in enumerate(sample_indices):
                    if idx < 50:  # Limit to 50 points
                        if server_name == list(server_cols.keys())[0]:  # First server, add time labels
                            ws.cell(row=data_row + idx, column=1).value = f"T+{idx}"
                        
                        cpu_value = round(cpu_data.iloc[data_idx], 1)
                        ws.cell(row=data_row + idx, column=server_cols[server_name]).value = cpu_value
                        
                        # Highlight anomaly cells
                        if data_idx in server_anomalies.get(server_name, []):
                            ws.cell(row=data_row + idx, column=server_cols[server_name]).fill = self.YELLOW_FILL
        
        # Second pass - add separate anomaly marker columns for EACH server
        # This allows us to distinguish which anomaly belongs to which server
        server_anomaly_cols = {}
        sample_indices = list(range(0, min(500, max([len(df) for df in all_data.values() if df is not None])), 10))
        
        for server_name in server_cols.keys():
            # Add column header for this server's anomalies
            ws.cell(row=chart_start_row, column=anomaly_line_start_col).value = f"{server_name} Anomalies"
            ws.cell(row=chart_start_row, column=anomaly_line_start_col).font = Font(bold=True, size=9)
            server_anomaly_cols[server_name] = anomaly_line_start_col
            anomaly_line_start_col += 1
            
            # Populate anomaly data for this specific server
            for idx, data_idx in enumerate(sample_indices):
                if idx < 50:
                    # Check if THIS server has an anomaly at this timestamp
                    if data_idx in server_anomalies.get(server_name, []):
                        # Set marker at 105% of max for visibility
                        ws.cell(row=data_row + idx, column=server_anomaly_cols[server_name]).value = max_cpu_value * 1.05
                    else:
                        # Set None so the series doesn't connect points
                        ws.cell(row=data_row + idx, column=server_anomaly_cols[server_name]).value = None
        
        # Create CPU chart with vertical anomaly lines
        if len(server_cols) > 0:
            chart = LineChart()
            chart.title = f"{location} - CPU Utilization Over Time (Distinct markers per device)"
            chart.style = 12
            chart.y_axis.title = "CPU %"
            chart.x_axis.title = "Time Intervals"
            chart.height = 10
            chart.width = 20
            
            # Add all data series (including anomaly marker columns)
            max_col = anomaly_line_start_col - 1  # Include all columns up to the last anomaly column
            data = Reference(ws, min_col=2, min_row=chart_start_row, max_row=min(data_row + 49, ws.max_row), max_col=max_col)
            cats = Reference(ws, min_col=1, min_row=chart_start_row + 1, max_row=min(data_row + 49, ws.max_row))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Style the series
            from openpyxl.chart.marker import Marker
            from openpyxl.drawing.line import LineProperties
            
            # Style normal server lines
            for idx in range(len(server_cols)):
                if idx < len(chart.series):
                    chart.series[idx].graphicalProperties.line.width = 25000  # 2.5pt
                    chart.series[idx].smooth = True
            
            # Style each server's anomaly marker series with distinct shapes and colors
            # Different marker shapes for each server's anomalies
            marker_shapes = ['triangle', 'diamond', 'circle', 'square', 'star', 'x', 'plus']
            marker_colors = ['FF0000', 'FF6600', 'CC0000', 'FF3333', '990000', 'FF9999', 'DD0000']
            
            num_servers = len(server_cols)
            for i in range(num_servers):
                anomaly_series_idx = num_servers + i
                if anomaly_series_idx < len(chart.series):
                    anomaly_series = chart.series[anomaly_series_idx]
                    # Assign unique marker shape and color for each server's anomalies
                    anomaly_series.marker = Marker(marker_shapes[i % len(marker_shapes)])
                    anomaly_series.marker.size = 10
                    anomaly_series.marker.graphicalProperties.solidFill = marker_colors[i % len(marker_colors)]
                    anomaly_series.marker.graphicalProperties.line.solidFill = marker_colors[i % len(marker_colors)]
                    # Remove the connecting line between markers
                    anomaly_series.graphicalProperties.line.noFill = True
            
            ws.add_chart(chart, f"K{stats_start_row}")
        
        # Create IOPS data and chart
        row = data_row + 52
        iops_start_row = row
        ws.cell(row=row, column=1).value = "Time Interval"
        col_idx = 2
        iops_cols = {}
        
        # First, add all server IOPS columns
        for server_name, df in all_data.items():
            if df is not None and not df.empty and 'Total IOPS' in df.columns:
                ws.cell(row=row, column=col_idx).value = f"{server_name} IOPS"
                iops_cols[server_name] = col_idx
                col_idx += 1
        
        # Add anomaly marker column
        iops_anomaly_line_col = col_idx
        ws.cell(row=row, column=col_idx).value = "IOPS Anomaly Marker"
        col_idx += 1
        
        # Collect all IOPS anomalies
        iops_server_anomalies = {}
        for server_name, df in all_data.items():
            if df is not None and not df.empty and 'Total IOPS' in df.columns:
                iops_anomaly_indices = self.detect_anomalies(df, 'Total IOPS')
                iops_server_anomalies[server_name] = iops_anomaly_indices
        
        # Write IOPS sample data
        data_row = row + 1
        max_iops_value = 0
        
        # First pass - write data and find max
        for server_name, df in all_data.items():
            if df is not None and not df.empty and 'Total IOPS' in df.columns and server_name in iops_cols:
                iops_data = df['Total IOPS'].dropna()
                max_iops_value = max(max_iops_value, iops_data.max())
                
                sample_indices = range(0, min(len(iops_data), 500), 10)
                
                for idx, data_idx in enumerate(sample_indices):
                    if idx < 50:
                        if server_name == list(iops_cols.keys())[0]:
                            ws.cell(row=data_row + idx, column=1).value = f"T+{idx}"
                        
                        iops_value = round(iops_data.iloc[data_idx], 0)
                        ws.cell(row=data_row + idx, column=iops_cols[server_name]).value = iops_value
                        
                        # Highlight anomaly
                        if data_idx in iops_server_anomalies.get(server_name, []):
                            ws.cell(row=data_row + idx, column=iops_cols[server_name]).fill = self.YELLOW_FILL
        
        # Second pass - add separate anomaly marker columns for EACH server (IOPS)
        iops_server_anomaly_cols = {}
        sample_indices = list(range(0, min(500, max([len(df) for df in all_data.values() if df is not None])), 10))
        
        for server_name in iops_cols.keys():
            # Add column header for this server's IOPS anomalies
            ws.cell(row=iops_start_row, column=iops_anomaly_line_col).value = f"{server_name} IOPS Anomalies"
            ws.cell(row=iops_start_row, column=iops_anomaly_line_col).font = Font(bold=True, size=9)
            iops_server_anomaly_cols[server_name] = iops_anomaly_line_col
            iops_anomaly_line_col += 1
            
            # Populate anomaly data for this specific server
            for idx, data_idx in enumerate(sample_indices):
                if idx < 50:
                    # Check if THIS server has an IOPS anomaly at this timestamp
                    if data_idx in iops_server_anomalies.get(server_name, []):
                        # Set marker at 105% of max for visibility
                        ws.cell(row=data_row + idx, column=iops_server_anomaly_cols[server_name]).value = max_iops_value * 1.05
                    else:
                        # Set None so the series doesn't connect points
                        ws.cell(row=data_row + idx, column=iops_server_anomaly_cols[server_name]).value = None
        
        # Create IOPS chart with vertical anomaly lines
        if len(iops_cols) > 0:
            chart2 = LineChart()
            chart2.title = f"{location} - IOPS Over Time (Distinct markers per device)"
            chart2.style = 13
            chart2.y_axis.title = "IOPS"
            chart2.x_axis.title = "Time Intervals"
            chart2.height = 10
            chart2.width = 20
            
            max_col2 = iops_anomaly_line_col - 1  # Include all columns up to the last IOPS anomaly column
            data2 = Reference(ws, min_col=2, min_row=iops_start_row, max_row=min(data_row + 49, ws.max_row), max_col=max_col2)
            cats2 = Reference(ws, min_col=1, min_row=iops_start_row + 1, max_row=min(data_row + 49, ws.max_row))
            
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)
            
            # Style the series
            from openpyxl.chart.marker import Marker
            
            # Style normal server lines
            for idx in range(len(iops_cols)):
                if idx < len(chart2.series):
                    chart2.series[idx].graphicalProperties.line.width = 25000  # 2.5pt
                    chart2.series[idx].smooth = True
            
            # Style each server's IOPS anomaly marker series with distinct shapes and colors
            marker_shapes = ['triangle', 'diamond', 'circle', 'square', 'star', 'x', 'plus']
            marker_colors = ['FF0000', 'FF6600', 'CC0000', 'FF3333', '990000', 'FF9999', 'DD0000']
            
            num_servers = len(iops_cols)
            for i in range(num_servers):
                anomaly_series_idx = num_servers + i
                if anomaly_series_idx < len(chart2.series):
                    anomaly_series = chart2.series[anomaly_series_idx]
                    # Assign unique marker shape and color for each server's IOPS anomalies
                    anomaly_series.marker = Marker(marker_shapes[i % len(marker_shapes)])
                    anomaly_series.marker.size = 10
                    anomaly_series.marker.graphicalProperties.solidFill = marker_colors[i % len(marker_colors)]
                    anomaly_series.marker.graphicalProperties.line.solidFill = marker_colors[i % len(marker_colors)]
                    # Remove the connecting line between markers
                    anomaly_series.graphicalProperties.line.noFill = True
            
            ws.add_chart(chart2, f"K{iops_start_row}")
        
        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
        
        return ws
    
    def create_capacity_chart(self, ws, location, start_row):
        """Create a pie chart showing capacity utilization."""
        chart = PieChart()
        chart.title = f"{location} - Storage Capacity Distribution"
        chart.style = 10
        chart.height = 10
        chart.width = 15
        
        # This is a placeholder - actual implementation would aggregate capacity data
        # and create proper chart references
        
        return chart
    
    def generate_location_report(self, location, servers):
        """Generate a complete report for one location."""
        print(f"Generating report for {location}...")
        
        wb = Workbook()
        # Remove the default empty sheet so we fully control sheet order/names
        wb.remove(wb.active)
        self.add_location_to_workbook(wb, location, servers)
        
        # Save the workbook
        output_file = os.path.join(self.output_dir, f'{location}_Storage_Report.xlsx')
        wb.save(output_file)
        print(f"  ✓ Saved: {output_file}")
        
        return output_file

    def add_location_to_workbook(self, wb, location, servers):
        """Add one location's report + analytics sheets into an existing workbook."""
        ws = wb.create_sheet(f"{location} Report")
        
        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18
        
        # Create main header
        current_row = self.create_header(ws, location, servers)
        
        # Collect all server data
        all_server_data = {}
        
        # Process each server
        for server in servers:
            df = None
            perf_stats = {}

            if server in self._server_data:
                df = self._server_data[server]
            elif self.raw_csv_dir and os.path.isdir(self.raw_csv_dir):
                csv_file = None
                for filename in os.listdir(self.raw_csv_dir):
                    if server in filename and filename.endswith('.csv'):
                        csv_file = os.path.join(self.raw_csv_dir, filename)
                        break
                if csv_file:
                    df = self.read_performance_data(csv_file)

            all_server_data[server] = df
            if df is not None and not df.empty:
                for metric in ['Latency', 'Avg. Size', 'Total IOPS', 'CPU Utilization']:
                    if metric in df.columns:
                        stats = self.calculate_statistics(df, metric)
                        if stats and 'mean' in stats and stats['mean'] > 0:
                            perf_stats[metric] = stats

            capacity_data = self._capacity_data.get(server, {})
            current_row = self.create_server_section(ws, server, capacity_data, perf_stats, current_row)
        
        # Create analytics sheet with charts (optional)
        if self.enable_analytics:
            has_data = any(df is not None and not df.empty for df in all_server_data.values())
            if has_data:
                self.create_charts_sheet(wb, location, all_server_data)
        
        return wb

    def generate_combined_report(self):
        """Generate one big Excel workbook containing all locations (in order)."""
        output_file = os.path.join(self.output_dir, 'All_Locations_Storage_Report.xlsx')
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Add all location reports (each adds its own report + analytics sheets)
        for location, servers in self.location_servers.items():
            self.add_location_to_workbook(wb, location, servers)
        
        wb.save(output_file)
        return output_file
    
    def generate_all_reports(self):
        """Generate reports for all locations."""
        print("=" * 70)
        print("Vodafone Storage Performance Report Generator")
        print("=" * 70)
        print()
        
        generated_files = []
        
        for location, servers in self.location_servers.items():
            try:
                output_file = self.generate_location_report(location, servers)
                generated_files.append(output_file)
            except Exception as e:
                print(f"  ✗ Error generating report for {location}: {e}")
                import traceback
                traceback.print_exc()

        # Also generate one combined workbook containing every location
        try:
            combined_file = self.generate_combined_report()
            generated_files.append(combined_file)
            print(f"\n  ✓ Saved combined report: {combined_file}")
        except Exception as e:
            print(f"\n  ✗ Error generating combined report: {e}")
            import traceback
            traceback.print_exc()
        
        print()
        print("=" * 70)
        print(f"Report Generation Complete! Generated {len(generated_files)} reports.")
        print("=" * 70)
        
        return generated_files


def main():
    """Main function."""
    generator = ReportGenerator()
    generated_files = generator.generate_all_reports()
    
    if generated_files:
        print("\nGenerated files:")
        for f in generated_files:
            print(f"  • {f}")


if __name__ == '__main__':
    main()

